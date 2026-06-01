"""
経済産業省 FIT/FIP認定情報 市町村別集計インポートスクリプト
FIT失効が多い市町村（=BESS/太陽光連携候補エリア）を特定するための基盤データを構築する。

ダウンロード手順:
  1. https://www.fit-portal.go.jp/publicinfosummary にアクセス
  2. 「B表 市町村別認定・導入量（最新版）」をダウンロード（Excel/CSV）
  3. このスクリプトを実行

使用方法:
  python -m app.scripts.import_fit_solar --file path/to/B表.xlsx
  python -m app.scripts.import_fit_solar --file path/to/B表.csv

ダウンロードせずにスクリプトをテスト:
  python -m app.scripts.import_fit_solar --demo
"""
import argparse
import sys
import os
from pathlib import Path

from app.database import SessionLocal, engine
from app import models

models.Base.metadata.create_all(bind=engine)

# FIT認定年度別の失効タイミング（20年期間）
# 2012年以前認定 → 2032年以前に失効
# 2012-2015年認定 → 2032-2035年失効
# 2016-2020年認定 → 2036-2040年失効

# B表のCSVカラム名マッピング（年度によって変わる可能性あり）
_PREF_COLS = ["都道府県名", "都道府県", "prefecture"]
_MUNI_COLS = ["市町村名", "市区町村名", "municipality", "市町村"]
_CERT_KW_COLS = ["認定容量（kW）", "認定容量", "certified_kw", "認定(kW)"]
_CERT_CNT_COLS = ["認定件数", "件数", "certified_count"]

# デモ用サンプルデータ（実データがない場合の動作確認用）
DEMO_DATA = [
    # (prefecture, municipality, certified_kw, certified_count)
    ("福島県", "福島市",   85000, 1200),
    ("福島県", "郡山市",  120000, 1800),
    ("福島県", "いわき市", 95000, 1400),
    ("福島県", "会津若松市", 35000, 500),
    ("宮城県", "仙台市",   75000, 1100),
    ("宮城県", "石巻市",   45000, 650),
    ("茨城県", "水戸市",   60000, 900),
    ("茨城県", "つくば市", 55000, 800),
]


def _get_col(row_dict: dict, candidates: list[str]):
    for c in candidates:
        if c in row_dict:
            return row_dict[c]
    return None


def import_csv(filepath: str, data_year: int = 2025) -> int:
    """CSVまたはExcelファイルからFIT認定情報をインポート"""
    ext = Path(filepath).suffix.lower()
    rows = []

    if ext in (".xls", ".xlsx"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, [v for v in row])))
        except ImportError:
            print("openpyxl が必要です: pip install openpyxl")
            sys.exit(1)
    elif ext == ".csv":
        import csv
        enc = "utf-8-sig"
        with open(filepath, encoding=enc, errors="replace") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
    else:
        print(f"未対応の形式: {ext}。.csv または .xlsx を使用してください")
        sys.exit(1)

    print(f"  {len(rows):,} 行読み込み")
    return _insert_rows(rows, data_year)


def import_demo(data_year: int = 2025) -> int:
    """デモデータでFIT認定情報をインポート（動作確認用）"""
    print("  デモデータを使用します（実データは fit-portal.go.jp からダウンロード）")
    rows = [
        {"都道府県名": p, "市町村名": m, "認定容量（kW）": kw, "認定件数": cnt}
        for p, m, kw, cnt in DEMO_DATA
    ]
    return _insert_rows(rows, data_year)


def _insert_rows(rows: list[dict], data_year: int) -> int:
    db = SessionLocal()
    try:
        # 同一年度のデータを削除
        existing = db.query(models.FitMunicipalitySolar).filter_by(data_year=data_year).count()
        if existing > 0:
            db.query(models.FitMunicipalitySolar).filter_by(data_year=data_year).delete()
            db.commit()
            print(f"  既存 {existing} 件を削除")

        count = 0
        skipped = 0
        for row in rows:
            pref = _get_col(row, _PREF_COLS)
            muni = _get_col(row, _MUNI_COLS)
            if not pref or not muni:
                skipped += 1
                continue

            try:
                kw = float(str(_get_col(row, _CERT_KW_COLS) or 0).replace(",", "") or 0)
                cnt = int(str(_get_col(row, _CERT_CNT_COLS) or 0).replace(",", "") or 0)
            except (ValueError, TypeError):
                skipped += 1
                continue

            if kw <= 0:
                continue

            db.add(models.FitMunicipalitySolar(
                prefecture=str(pref).strip(),
                municipality=str(muni).strip(),
                certified_kw=kw,
                certified_count=cnt,
                data_year=data_year,
            ))
            count += 1

            if count % 5000 == 0:
                db.commit()
                print(f"  {count:,} 件インポート中...")

        db.commit()
        print(f"  完了: {count:,} 件インポート（{skipped} 件スキップ）")
        return count
    finally:
        db.close()


def run():
    parser = argparse.ArgumentParser(description="FIT認定情報（市町村別）インポーター")
    parser.add_argument("--file", help="CSVまたはExcelファイルのパス")
    parser.add_argument("--demo", action="store_true", help="デモデータで動作確認")
    parser.add_argument("--year", type=int, default=2025, help="データ年度（デフォルト: 2025）")
    args = parser.parse_args()

    if args.demo:
        print("=== FIT認定情報デモインポート ===")
        n = import_demo(args.year)
    elif args.file:
        if not os.path.exists(args.file):
            print(f"ファイルが見つかりません: {args.file}")
            sys.exit(1)
        print(f"=== FIT認定情報インポート: {args.file} ===")
        n = import_csv(args.file, args.year)
    else:
        print("使用方法:")
        print("  --demo : デモデータで動作確認")
        print("  --file <path> : CSVまたはExcelファイルのパス")
        print("")
        print("データは https://www.fit-portal.go.jp/publicinfosummary から")
        print("「B表 市町村別認定・導入量」をダウンロードして使用してください")
        return

    print(f"\n合計 {n:,} 市町村レコードをインポートしました")


if __name__ == "__main__":
    run()
